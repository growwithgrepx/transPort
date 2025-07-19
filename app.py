from dotenv import load_dotenv

load_dotenv()
from flask import Flask, render_template, redirect, url_for, request, flash, session, jsonify, make_response, send_file
from extensions import db
from werkzeug.security import generate_password_hash, check_password_hash
import os
import re
import json
from datetime import datetime
from flask_migrate import Migrate
import click
from flask.cli import with_appcontext
from flask_login import LoginManager, login_required, current_user, login_user, logout_user
from wtforms import Form, StringField, PasswordField, validators
from wtforms.validators import DataRequired, Email, Length, Optional
from math import ceil
from flask_admin import Admin, expose
from flask_admin.contrib.sqla import ModelView
from flask import abort
from flask_wtf.csrf import CSRFProtect, generate_csrf
from models import User
import logging
from logging.handlers import RotatingFileHandler
import sentry_sdk
from sentry_sdk.integrations.flask import FlaskIntegration
from sentry_sdk.integrations.sqlalchemy import SqlalchemyIntegration
import traceback
from functools import wraps
import io
import pandas as pd
from models import Driver, Agent, Vehicle, Service, Billing, Discount, Job
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

app = Flask(__name__)

# Initialize Sentry for error tracking
sentry_dsn = os.environ.get('SENTRY_DSN')
if sentry_dsn:
    sentry_sdk.init(
        dsn=sentry_dsn,
        integrations=[
            FlaskIntegration(),
            SqlalchemyIntegration(),
        ],
        traces_sample_rate=1.0,
        environment=os.environ.get('FLASK_ENV', 'development')
    )

# Configure logging
if not app.debug:
    if not os.path.exists('logs'):
        os.mkdir('logs')

    file_handler = RotatingFileHandler('logs/transport_app.log', maxBytes=10240, backupCount=10)
    file_handler.setFormatter(logging.Formatter(
        '%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]'
    ))
    file_handler.setLevel(logging.INFO)
    app.logger.addHandler(file_handler)
    app.logger.setLevel(logging.INFO)
    app.logger.info('Transport Admin Portal startup')

# # Disable CSRF in development
# if app.config['DEBUG']:
#     app.config['WTF_CSRF_ENABLED'] = False


class Config:
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    SECRET_KEY = os.environ.get('SECRET_KEY', 'your-secret-key-change-this-in-production')
    db_url = os.environ.get('DATABASE_URL', '')
    if db_url.startswith('postgres://'):
        db_url = db_url.replace('postgres://', 'postgresql://')
    if not db_url:
        db_url = 'sqlite:///app.db'
    SQLALCHEMY_DATABASE_URI = db_url


class DevelopmentConfig(Config):
    pass


class TestingConfig(Config):
    TESTING = True


class ProductionConfig(Config):
    pass


config_map = {
    'development': DevelopmentConfig,
    'testing': TestingConfig,
    'production': ProductionConfig
}

app_env = os.environ.get('FLASK_ENV', 'development')
app.config.from_object(config_map.get(app_env, DevelopmentConfig))

if not app.config['SQLALCHEMY_DATABASE_URI']:
    raise RuntimeError('DATABASE_URL environment variable must be set to a valid PostgreSQL connection string.')

db.init_app(app)
csrf = CSRFProtect(app)

# Flask-Login setup
login_manager = LoginManager()
login_manager.init_app(app)
setattr(login_manager, 'login_view', 'login')
login_manager.login_message = 'Please log in to access this page.'

migrate = Migrate(app, db)

with app.app_context():
    from models import User, Job, Driver, Agent, Billing, Discount, Service, Vehicle, Role
from services.billing_service import BillingService


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


class LoginForm(Form):
    username = StringField('Username or Email', validators=[
        DataRequired(message='Username or email is required'),
        Length(min=3, max=150, message='Username must be between 3 and 150 characters')
    ])
    password = PasswordField('Password', validators=[
        DataRequired(message='Password is required'),
        Length(min=3, max=255, message='Password must be between 3 and 255 characters')
    ])


# Input validation decorators
def validate_json_input(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        try:
            if request.is_json:
                data = request.get_json()
                if not data:
                    app.logger.warning(f'Empty JSON data received in {f.__name__}')
                    return jsonify({'error': 'No input data provided'}), 400
            return f(*args, **kwargs)
        except Exception as e:
            app.logger.error(f'JSON validation error in {f.__name__}: {str(e)}')
            return jsonify({'error': 'Invalid JSON data'}), 400

    return decorated_function


def validate_form_input(required_fields=None):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if required_fields and request.method == 'POST':
                for field in required_fields:
                    if not request.form.get(field):
                        app.logger.warning(f'Missing required field {field} in {f.__name__}')
                        flash(f'{field.replace("_", " ").title()} is required', 'error')
                        return redirect(request.url)
            return f(*args, **kwargs)

        return decorated_function

    return decorator


def handle_database_errors(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        try:
            return f(*args, **kwargs)
        except Exception as e:
            db.session.rollback()
            app.logger.error(f'Database error in {f.__name__}: {str(e)}')
            app.logger.error(traceback.format_exc())
            flash('An error occurred while processing your request. Please try again.', 'error')
            return redirect(url_for('dashboard'))

    return decorated_function


# Custom admin view to restrict access to admins only
class AdminModelView(ModelView):
    def is_accessible(self):
        return current_user.is_authenticated and any(
            role.name in ['fleet_manager', 'system_admin'] for role in current_user.roles)

    def inaccessible_callback(self, name, **kwargs):
        return abort(403)


# Initialize Flask-Admin
admin = Admin(app, name='Admin', template_mode='bootstrap4')

with app.app_context():
    admin.add_view(AdminModelView(User, db.session))
    admin.add_view(AdminModelView(Role, db.session))
    admin.add_view(AdminModelView(Job, db.session))
    admin.add_view(AdminModelView(Driver, db.session))
    admin.add_view(AdminModelView(Agent, db.session))
    admin.add_view(AdminModelView(Vehicle, db.session))
    admin.add_view(AdminModelView(Billing, db.session))
    admin.add_view(AdminModelView(Discount, db.session))
    admin.add_view(AdminModelView(Service, db.session))


# Error handlers
@app.errorhandler(400)
def bad_request(error):
    app.logger.error(f'Bad request: {error}')
    return render_template('errors/400.html'), 400


@app.errorhandler(403)
def forbidden(error):
    app.logger.error(f'Forbidden access: {error}')
    return render_template('errors/403.html'), 403


@app.errorhandler(404)
def not_found(error):
    app.logger.error(f'Page not found: {error}')
    return render_template('errors/404.html'), 404


@app.errorhandler(500)
def internal_error(error):
    app.logger.error(f'Internal server error: {error}')
    app.logger.error(traceback.format_exc())
    db.session.rollback()
    return render_template('errors/500.html'), 500


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        app.logger.info(f'User {current_user.username} already authenticated, redirecting to dashboard')
        return redirect(url_for('dashboard'))

    form = LoginForm(request.form)
    error = None

    if request.method == 'POST':
        app.logger.info(f'Login attempt for user: {request.form.get("username", "unknown")}')

        if form.validate():
            username = (form.username.data or '').strip()
            password = form.password.data or ''

            # Input sanitization
            if not username or not password:
                error = 'Username and password are required'
                app.logger.warning('Empty username or password provided')
            else:
                try:
                    # Try to find user by username or email
                    user = User.query.filter(
                        (User.username == username) | (User.email == username)
                    ).first()

                    if user and user.check_password(password):
                        if user.active:
                            login_user(user)
                            app.logger.info(f'User {user.username} logged in successfully')
                            flash('Login successful!', 'success')
                            next_page = request.args.get('next')
                            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
                        else:
                            error = 'Account is inactive. Please contact administrator.'
                            app.logger.warning(f'Inactive user {user.username} attempted to login')
                    else:
                        error = 'Invalid username or password'
                        app.logger.warning(f'Failed login attempt for user: {username}')

                except Exception as e:
                    app.logger.error(f'Login error: {str(e)}')
                    error = 'An error occurred during login. Please try again.'
        else:
            error = 'Please correct the errors below'
            app.logger.warning(f'Form validation failed: {form.errors}')

    return render_template('login.html', form=form, error=error)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    return redirect(url_for('dashboard'))

@app.route('/dashboard')
@login_required
def dashboard():
    from models import Job, Vehicle, Driver
    from datetime import date
    # Unassigned Jobs: jobs with no driver or vehicle assigned
    unassigned_jobs = Job.query.filter((Job.driver_id == None) | (Job.vehicle_type == None)).count()
    # Ready to Invoice: jobs with order_status 'Completed' and not yet billed (assuming payment_status 'Unpaid')
    ready_to_invoice = Job.query.filter(Job.order_status == 'Completed', Job.payment_status == 'Unpaid').count()
    # Total Vehicles
    total_vehicles = Vehicle.query.count()
    # Available Drivers: drivers not assigned to any active job (assuming jobs with order_status 'New' or 'In Progress')
    active_driver_ids = [job.driver_id for job in Job.query.filter(Job.order_status.in_(['New', 'In Progress'])).all()
                         if job.driver_id]
    available_drivers = Driver.query.filter(~Driver.id.in_(active_driver_ids)).count()
    # Active Jobs: jobs with order_status 'New' or 'In Progress'
    active_jobs = Job.query.filter(Job.order_status.in_(['New', 'In Progress'])).count()
    # Completed Today: jobs with order_status 'Completed' and pickup_date is today
    completed_today = Job.query.filter(Job.order_status == 'Completed',
                                       Job.pickup_date == date.today().isoformat()).count()

   
    return render_template(
        'dashboard.html',
        unassigned_jobs=unassigned_jobs,
        ready_to_invoice=ready_to_invoice,
        total_vehicles=total_vehicles,
        available_drivers=available_drivers,
        active_jobs=active_jobs,
        completed_today=completed_today,
    )

@app.route('/expriment')
@login_required
def expriment():
     return render_template('expriment.html' )


@app.route('/demo')
@login_required
def demo():
    from models import Job, Vehicle, Driver
    from datetime import date
    # Unassigned Jobs: jobs with no driver or vehicle assigned
    unassigned_jobs = Job.query.filter((Job.driver_id == None) | (Job.vehicle_type == None)).count()
    # Ready to Invoice: jobs with order_status 'Completed' and not yet billed (assuming payment_status 'Unpaid')
    ready_to_invoice = Job.query.filter(Job.order_status == 'Completed', Job.payment_status == 'Unpaid').count()
    # Total Vehicles
    total_vehicles = Vehicle.query.count()
    # Available Drivers: drivers not assigned to any active job (assuming jobs with order_status 'New' or 'In Progress')
    active_driver_ids = [job.driver_id for job in Job.query.filter(Job.order_status.in_(['New', 'In Progress'])).all()
                         if job.driver_id]
    available_drivers = Driver.query.filter(~Driver.id.in_(active_driver_ids)).count()
    # Active Jobs: jobs with order_status 'New' or 'In Progress'
    active_jobs = Job.query.filter(Job.order_status.in_(['New', 'In Progress'])).count()
    # Completed Today: jobs with order_status 'Completed' and pickup_date is today
    completed_today = Job.query.filter(Job.order_status == 'Completed',
                                       Job.pickup_date == date.today().isoformat()).count()
    return render_template('demo.html',
                           unassigned_jobs=unassigned_jobs,
                           ready_to_invoice=ready_to_invoice,
                           total_vehicles=total_vehicles,
                           available_drivers=available_drivers,
                           active_jobs=active_jobs,
                           completed_today=completed_today
                           )


# JOBS CRUD
@app.route('/jobs', methods=['GET', 'POST'])
@login_required
def jobs():
    page = request.args.get('page', 1, type=int)
    per_page = 20
    search_fields = [
        'customer_name', 'customer_email', 'customer_mobile', 'customer_reference',
        'passenger_name', 'passenger_email', 'passenger_mobile', 'type_of_service',
        'pickup_date', 'pickup_time', 'pickup_location', 'dropoff_location',
        'vehicle_type', 'vehicle_number', 'driver_contact', 'payment_mode',
        'payment_status', 'order_status', 'message', 'remarks', 'reference', 'status'
    ]
    filters = []
    advanced = False
    for field in search_fields:
        value = request.args.get(field)
        if value:
            advanced = True
            filters.append(getattr(Job, field).ilike(f'%{value}%'))
    search_query = request.args.get('search', '')
    if advanced:
        query = Job.query.filter(*filters)
    elif search_query:
        query = Job.query.filter(
            (Job.customer_name.ilike(f'%{search_query}%')) |
            (Job.customer_email.ilike(f'%{search_query}%')) |
            (Job.customer_mobile.ilike(f'%{search_query}%')) |
            (Job.customer_reference.ilike(f'%{search_query}%')) |
            (Job.passenger_name.ilike(f'%{search_query}%')) |
            (Job.passenger_email.ilike(f'%{search_query}%')) |
            (Job.passenger_mobile.ilike(f'%{search_query}%')) |
            (Job.type_of_service.ilike(f'%{search_query}%')) |
            (Job.pickup_date.ilike(f'%{search_query}%')) |
            (Job.pickup_time.ilike(f'%{search_query}%')) |
            (Job.pickup_location.ilike(f'%{search_query}%')) |
            (Job.dropoff_location.ilike(f'%{search_query}%')) |
            (Job.vehicle_type.ilike(f'%{search_query}%')) |
            (Job.vehicle_number.ilike(f'%{search_query}%')) |
            (Job.driver_contact.ilike(f'%{search_query}%')) |
            (Job.payment_mode.ilike(f'%{search_query}%')) |
            (Job.payment_status.ilike(f'%{search_query}%')) |
            (Job.order_status.ilike(f'%{search_query}%')) |
            (Job.message.ilike(f'%{search_query}%')) |
            (Job.remarks.ilike(f'%{search_query}%')) |
            (Job.reference.ilike(f'%{search_query}%')) |
            (Job.status.ilike(f'%{search_query}%'))
        )
    else:
        query = Job.query
    pagination = query.order_by(Job.id.desc()).paginate(page=page, per_page=per_page, error_out=False)
    jobs = pagination.items
    return render_template('jobs.html', jobs=jobs, search_query=search_query, pagination=pagination)


@app.route('/jobs/table', methods=['GET'])
@login_required
def jobs_table():
    page = request.args.get('page', 1, type=int)
    per_page = 20
    search_fields = [
        'customer_name', 'customer_email', 'customer_mobile', 'customer_reference',
        'passenger_name', 'passenger_email', 'passenger_mobile', 'type_of_service',
        'pickup_date', 'pickup_time', 'pickup_location', 'dropoff_location',
        'vehicle_type', 'vehicle_number', 'driver_contact', 'payment_mode',
        'payment_status', 'order_status', 'message', 'remarks', 'reference', 'status'
    ]
    filters = []
    advanced = False
    for field in search_fields:
        value = request.args.get(field)
        if value:
            advanced = True
            filters.append(getattr(Job, field).ilike(f'%{value}%'))
    search_query = request.args.get('search', '')
    if advanced:
        query = Job.query.filter(*filters)
    elif search_query:
        query = Job.query.filter(
            (Job.customer_name.ilike(f'%{search_query}%')) |
            (Job.customer_email.ilike(f'%{search_query}%')) |
            (Job.customer_mobile.ilike(f'%{search_query}%')) |
            (Job.customer_reference.ilike(f'%{search_query}%')) |
            (Job.passenger_name.ilike(f'%{search_query}%')) |
            (Job.passenger_email.ilike(f'%{search_query}%')) |
            (Job.passenger_mobile.ilike(f'%{search_query}%')) |
            (Job.type_of_service.ilike(f'%{search_query}%')) |
            (Job.pickup_date.ilike(f'%{search_query}%')) |
            (Job.pickup_time.ilike(f'%{search_query}%')) |
            (Job.pickup_location.ilike(f'%{search_query}%')) |
            (Job.dropoff_location.ilike(f'%{search_query}%')) |
            (Job.vehicle_type.ilike(f'%{search_query}%')) |
            (Job.vehicle_number.ilike(f'%{search_query}%')) |
            (Job.driver_contact.ilike(f'%{search_query}%')) |
            (Job.payment_mode.ilike(f'%{search_query}%')) |
            (Job.payment_status.ilike(f'%{search_query}%')) |
            (Job.order_status.ilike(f'%{search_query}%')) |
            (Job.message.ilike(f'%{search_query}%')) |
            (Job.remarks.ilike(f'%{search_query}%')) |
            (Job.reference.ilike(f'%{search_query}%')) |
            (Job.status.ilike(f'%{search_query}%'))
        )
    else:
        query = Job.query
    pagination = query.order_by(Job.id.desc()).paginate(page=page, per_page=per_page, error_out=False)
    jobs = pagination.items
    return render_template('jobs_table.html', jobs=jobs, pagination=pagination)


@app.route('/jobs/add', methods=['GET', 'POST'])
# @login_required
@handle_database_errors
def add_job():
    from models import Agent, Service, Vehicle, Driver
    agents = Agent.query.filter_by(status='Active').all()
    services = Service.query.filter_by(status='Active').all()
    vehicles = Vehicle.query.filter_by(status='Active').all()
    drivers = Driver.query.all()
    
    if request.method == 'POST':
            
            return handle_single_job_creation()
    
    now = datetime.now()
    return render_template('view_job.html', job=None, agents=agents, services=services, vehicles=vehicles,
                           drivers=drivers, current_date=now.strftime('%Y-%m-%d'),current_time=now.strftime('%H:%M'))


@app.route('/jobs/add_bulk', methods=['GET', 'POST'])
@login_required
@handle_database_errors
def add_bulk_jobs():
    from models import Agent, Service, Vehicle, Driver
    agents = Agent.query.filter_by(status='Active').all()
    services = Service.query.filter_by(status='Active').all()
    vehicles = Vehicle.query.filter_by(status='Active').all()
    drivers = Driver.query.all()
    
    if request.method == 'POST':
        return handle_bulk_job_creation()
    
    return render_template('bulk_jobs.html', agents=agents, services=services, vehicles=vehicles,
                           drivers=drivers)


def handle_single_job_creation():
    """Handle single job creation with validation and form data preservation"""
    try:
        form_data = request.form.to_dict()
        print(form_data)
        errors = {}
        print("handle single job creation 1")
        agent_id = request.form.get('agent_id')
        service_id = request.form.get('service_id')
        vehicle_id = request.form.get('vehicle_id')
        driver_id = request.form.get('driver_id')

        agent = Agent.query.get(agent_id) if agent_id and agent_id.isdigit() else None
        service = Service.query.get(service_id) if service_id and service_id.isdigit() else None
        vehicle = Vehicle.query.get(vehicle_id) if vehicle_id and vehicle_id.isdigit() else None
        driver = Driver.query.get(driver_id) if driver_id and driver_id.isdigit() else None

        customer_name = agent.name if agent else request.form.get('customer_name', '').strip()
        pickup_location = request.form.get('pickup_location', '').strip()
        dropoff_location = request.form.get('dropoff_location', '').strip()
        pickup_date = request.form.get('pickup_date', '').strip()
        pickup_time = request.form.get('pickup_time', '').strip()

        customer_email = agent.email if agent else request.form.get('customer_email', '').strip()
        customer_mobile = agent.mobile if agent else request.form.get('customer_mobile', '').strip()
        passenger_email = request.form.get('passenger_email', '').strip()
        passenger_mobile = request.form.get('passenger_mobile', '').strip()


        if not customer_name:
            flash('Customer name is required', 'error')
            return redirect(request.url)
        if not pickup_location:
            flash('Pickup location is required', 'error')
            return redirect(request.url)
        if not dropoff_location:
            flash('Dropoff location is required', 'error')
            return redirect(request.url)
        if not pickup_date:
            flash('Pickup date is required', 'error')
            return redirect(request.url)

        # Validate pickup_date
        try:
            datetime.strptime(pickup_date, '%Y-%m-%d')
        except ValueError:
            flash('Invalid pickup date format', 'error')
            return redirect(request.url)

        # Optional: Validate pickup_time format (if provided)
        if pickup_time:
            try:
                datetime.strptime(pickup_time, '%H:%M')
            except ValueError:
                flash('Invalid pickup time format. Expected HH:MM.', 'error')
                return redirect(request.url)

        # Validate emails
        if passenger_email and not re.match(r'^[^@]+@[^@]+\.[^@]+$', passenger_email):
            errors['passenger_email'] = 'Invalid email format'
        if customer_email and not re.match(r'^[^@]+@[^@]+\.[^@]+$', customer_email):
            flash('Invalid customer email format', 'error')
            return redirect(request.url)

        # Validate mobile numbers
        if passenger_mobile and not re.match(r'^[\d\s\-\+\(\)]+$', passenger_mobile):
            errors['passenger_mobile'] = 'Invalid mobile number format'
        if customer_mobile and not re.match(r'^[\d\s\-\+\(\)]+$', customer_mobile):
            flash('Invalid customer mobile number format', 'error')
            return redirect(request.url)

        # If errors exist, re-render with preserved data and field-level errors
        print("handle single job creation 2 working ")
        if errors:
            print("handle single job creation 3 working ")
            return render_template('view_job.html',
                                   job=None,
                                   agents=Agent.query.filter_by(status='Active').all(),
                                   services=Service.query.filter_by(status='Active').all(),
                                   vehicles=Vehicle.query.filter_by(status='Active').all(),
                                   drivers=Driver.query.all(),
                                   errors=errors,
                                   form_data=form_data)

        # Collect pricing and other optional fields
        print("handle single job creation 3 working ", errors)
        base_price = float(request.form.get('base_price', 0) or 0)
        base_discount_percent = float(request.form.get('base_discount_percent', 0) or 0)
        agent_discount_percent = float(request.form.get('agent_discount_percent', 0) or 0)
        additional_discount_percent = float(request.form.get('additional_discount_percent', 0) or 0)
        additional_charges = float(request.form.get('additional_charges', 0) or 0)
        final_price = float(request.form.get('final_price', 0) or 0)
        invoice_number = request.form.get('invoice_number', '').strip()
        stops = request.form.getlist('additional_stops[]')

        job = Job(
            customer_name=customer_name,
            customer_email=customer_email,
            customer_mobile=customer_mobile,
            agent_id=agent.id if agent else None,
            type_of_service=service.name if service else request.form.get('type_of_service', '').strip(),
            vehicle_type=vehicle.type if vehicle else request.form.get('vehicle_type', '').strip(),
            vehicle_number=vehicle.number if vehicle else request.form.get('vehicle_number', '').strip(),
            driver_contact=driver.name if driver else request.form.get('driver_contact', '').strip(),
            driver_id=driver.id if driver else None,
            customer_reference=request.form.get('customer_reference', '').strip(),
            passenger_name=request.form.get('passenger_name', '').strip(),
            passenger_email=passenger_email,
            passenger_mobile=passenger_mobile,
            pickup_date=pickup_date,
            pickup_time=pickup_time,
            pickup_location=pickup_location,
            dropoff_location=dropoff_location,
            payment_mode=request.form.get('payment_mode', '').strip(),
            payment_status=request.form.get('payment_status', '').strip(),
            order_status=request.form.get('order_status', '').strip(),
            message=request.form.get('message', '').strip(),
            remarks=request.form.get('remarks', '').strip(),
            has_additional_stop=bool(request.form.get('has_additional_stop')),
            additional_stops=json.dumps(stops) if stops else None,
            has_request=bool(request.form.get('has_request')),
            reference=request.form.get('reference', '').strip(),
            status=request.form.get('status', '').strip(),
            date=pickup_date,
            # Pricing
            base_price=base_price,
            base_discount_percent=base_discount_percent,
            agent_discount_percent=agent_discount_percent,
            additional_discount_percent=additional_discount_percent,
            additional_charges=additional_charges,
            final_price=final_price,
            invoice_number=invoice_number
        )

        db.session.add(job)
        db.session.commit()
        app.logger.info(f'Job created successfully by user {current_user.username}: {job.id}')
        flash('Job created successfully', 'success')
        return redirect(url_for('jobs'))

    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error creating job: {str(e)}')
        flash('Error creating job. Please try again.', 'error')
        return redirect(request.url)


def handle_bulk_job_creation():
    """Handle bulk job creation"""
    try:
        jobs_data = request.form.get('jobs')
        if not jobs_data:
            flash('No jobs data received', 'error')
            return redirect(request.url)
        
        # Parse jobs data from form
        jobs = []
        for key, value in request.form.items():
            if key.startswith('jobs[') and key.endswith('][agent_id]'):
                # Extract row number from key like "jobs[1][agent_id]"
                row_num = key.split('[')[1].split(']')[0]
                
                # Get all data for this row
                agent_id = request.form.get(f'jobs[{row_num}][agent_id]', '').strip()
                service_id = request.form.get(f'jobs[{row_num}][service_id]', '').strip()
                vehicle_id = request.form.get(f'jobs[{row_num}][vehicle_id]', '').strip()
                driver_id = request.form.get(f'jobs[{row_num}][driver_id]', '').strip()
                pickup_date = request.form.get(f'jobs[{row_num}][pickup_date]', '').strip()
                pickup_time = request.form.get(f'jobs[{row_num}][pickup_time]', '').strip()
                pickup_location = request.form.get(f'jobs[{row_num}][pickup_location]', '').strip()
                dropoff_location = request.form.get(f'jobs[{row_num}][dropoff_location]', '').strip()
                passenger_name = request.form.get(f'jobs[{row_num}][passenger_name]', '').strip()
                
                # Validate required fields
                if not agent_id or not service_id or not vehicle_id or not driver_id or not pickup_date or not pickup_location or not dropoff_location:
                    flash(f'Row {row_num}: All required fields must be filled', 'error')
                    return redirect(request.url)
                
                # Get related objects
                agent = Agent.query.get(agent_id) if agent_id and agent_id.isdigit() else None
                service = Service.query.get(service_id) if service_id and service_id.isdigit() else None
                vehicle = Vehicle.query.get(vehicle_id) if vehicle_id and vehicle_id.isdigit() else None
                driver = Driver.query.get(driver_id) if driver_id and driver_id.isdigit() else None
                
                if not agent or not service or not vehicle or not driver:
                    flash(f'Row {row_num}: Invalid agent, service, vehicle, or driver selection', 'error')
                    return redirect(request.url)
                
                # Validate date format
                try:
                    datetime.strptime(pickup_date, '%Y-%m-%d')
                except ValueError:
                    flash(f'Row {row_num}: Invalid pickup date format', 'error')
                    return redirect(request.url)
                
                jobs.append({
                    'agent': agent,
                    'service': service,
                    'vehicle': vehicle,
                    'driver': driver,
                    'pickup_date': pickup_date,
                    'pickup_time': pickup_time,
                    'pickup_location': pickup_location,
                    'dropoff_location': dropoff_location,
                    'passenger_name': passenger_name
                })
        
        if not jobs:
            flash('No valid jobs to create', 'error')
            return redirect(request.url)
        
        # Create all jobs
        created_jobs = []
        for job_data in jobs:
            # Calculate pricing for each job
            base_price = job_data['service'].base_price or 0
            base_discount = Discount.query.filter_by(is_base_discount=True, is_active=True).first()
            base_discount_percent = (base_discount.percent if base_discount else 0.0) or 0.0
            agent_discount_percent = (job_data['agent'].agent_discount_percent if job_data['agent'] else 0.0) or 0.0
            
            # Calculate final price
            base_discount_amount = (base_price * base_discount_percent) / 100
            agent_discount_amount = (base_price * agent_discount_percent) / 100
            subtotal = base_price - base_discount_amount - agent_discount_amount
            final_price = subtotal
            
            job = Job(
                customer_name=job_data['agent'].name,
                customer_email=job_data['agent'].email,
                customer_mobile=job_data['agent'].mobile,
                agent_id=job_data['agent'].id,
                type_of_service=job_data['service'].name,
                vehicle_type=job_data['vehicle'].type,
                vehicle_number=job_data['vehicle'].number,
                driver_contact=job_data['driver'].name,
                driver_id=job_data['driver'].id,
                passenger_name=job_data['passenger_name'],
                pickup_date=job_data['pickup_date'],
                pickup_time=job_data['pickup_time'],
                pickup_location=job_data['pickup_location'],
                dropoff_location=job_data['dropoff_location'],
                status='Scheduled',
                date=job_data['pickup_date'],
                # Pricing fields
                base_price=base_price,
                base_discount_percent=base_discount_percent,
                agent_discount_percent=agent_discount_percent,
                additional_discount_percent=0.0,
                additional_charges=0.0,
                final_price=final_price
            )
            db.session.add(job)
            created_jobs.append(job)
        
        db.session.commit()
        
        app.logger.info(f'{len(created_jobs)} jobs created successfully by user {current_user.username}')
        flash(f'{len(created_jobs)} jobs created successfully!', 'success')
        return redirect(url_for('jobs'))
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error creating bulk jobs: {str(e)}')
        flash('Error creating jobs. Please try again.', 'error')
        return redirect(request.url)


@app.route('/jobs/edit/<int:job_id>', methods=['GET', 'POST'])
@login_required
def edit_job(job_id):
    from models import Agent, Service, Vehicle, Driver
    agents = Agent.query.filter_by(status='Active').all()
    services = Service.query.filter_by(status='Active').all()
    vehicles = Vehicle.query.filter_by(status='Active').all()
    drivers = Driver.query.all()
    job = Job.query.get_or_404(job_id)
    stops = json.loads(job.additional_stops) if job.additional_stops else []
    if request.method == 'POST':
        agent_id = request.form.get('agent_id')
        agent = Agent.query.get(agent_id) if agent_id else None
        service_id = request.form.get('service_id')
        service = Service.query.get(service_id) if service_id else None
        vehicle_id = request.form.get('vehicle_id')
        vehicle = Vehicle.query.get(vehicle_id) if vehicle_id else None
        driver_id = request.form.get('driver_id')
        driver = Driver.query.get(driver_id) if driver_id else None
        stops = request.form.getlist('additional_stops[]')
        job.customer_name = agent.name if agent else request.form.get('customer_name')
        job.customer_email = agent.email if agent else request.form.get('customer_email')
        job.customer_mobile = agent.mobile if agent else request.form.get('customer_mobile')
        job.agent_id = agent.id if agent else None
        job.type_of_service = service.name if service else request.form.get('type_of_service')
        job.vehicle_type = vehicle.type if vehicle else request.form.get('vehicle_type')
        job.vehicle_number = vehicle.number if vehicle else request.form.get('vehicle_number')
        job.driver_contact = driver.name if driver else request.form.get('driver_contact')
        job.driver_id = driver.id if driver else None
        job.customer_reference = request.form.get('customer_reference')
        job.passenger_name = request.form.get('passenger_name')
        job.passenger_email = request.form.get('passenger_email')
        job.passenger_mobile = request.form.get('passenger_mobile')
        job.pickup_date = request.form.get('pickup_date')
        job.pickup_time = request.form.get('pickup_time')
        job.pickup_location = request.form.get('pickup_location')
        job.dropoff_location = request.form.get('dropoff_location')
        job.payment_mode = request.form.get('payment_mode')
        job.payment_status = request.form.get('payment_status')
        job.order_status = request.form.get('order_status')
        job.message = request.form.get('message')
        job.remarks = request.form.get('remarks')
        job.has_additional_stop = bool(request.form.get('has_additional_stop'))
        job.additional_stops = json.dumps(stops) if stops else None
        job.has_request = bool(request.form.get('has_request'))
        job.reference = request.form.get('reference')
        job.status = request.form.get('status')
        job.date = request.form.get('pickup_date')
        
        # Update pricing fields
        job.base_price = float(request.form.get('base_price', 0) or 0)
        job.base_discount_percent = float(request.form.get('base_discount_percent', 0) or 0)
        job.agent_discount_percent = float(request.form.get('agent_discount_percent', 0) or 0)
        job.additional_discount_percent = float(request.form.get('additional_discount_percent', 0) or 0)
        job.additional_charges = float(request.form.get('additional_charges', 0) or 0)
        job.final_price = float(request.form.get('final_price', 0) or 0)
        job.invoice_number = request.form.get('invoice_number', '').strip()
        
        db.session.commit()
        return redirect(url_for('jobs'))
    return render_template('view_job.html', job=job, agents=agents, services=services, vehicles=vehicles,
                           drivers=drivers)


@app.route('/jobs/view/<int:job_id>', methods=['GET'])
@login_required
def view_job(job_id):
    job = Job.query.get_or_404(job_id)
    return render_template('view_job.html', job=job)


@app.route('/jobs/view/<int:job_id>/update', methods=['POST'])
@login_required
def update_job_view(job_id):
    job = Job.query.get_or_404(job_id)
    
    # Update job fields from form data
    job.customer_name = request.form.get('customer_name', job.customer_name)
    job.customer_email = request.form.get('customer_email', job.customer_email)
    job.customer_mobile = request.form.get('customer_mobile', job.customer_mobile)
    job.customer_reference = request.form.get('customer_reference', job.customer_reference)
    job.type_of_service = request.form.get('type_of_service', job.type_of_service)
    job.passenger_name = request.form.get('passenger_name', job.passenger_name)
    job.passenger_email = request.form.get('passenger_email', job.passenger_email)
    job.passenger_mobile = request.form.get('passenger_mobile', job.passenger_mobile)
    job.pickup_date = request.form.get('pickup_date', job.pickup_date)
    job.pickup_time = request.form.get('pickup_time', job.pickup_time)
    job.pickup_location = request.form.get('pickup_location', job.pickup_location)
    job.dropoff_location = request.form.get('dropoff_location', job.dropoff_location)
    job.status = request.form.get('status', job.status)
    job.payment_status = request.form.get('payment_status', job.payment_status)
    job.message = request.form.get('message', job.message)
    job.remarks = request.form.get('remarks', job.remarks)
    
    # Update pricing fields
    job.base_price = float(request.form.get('base_price', job.base_price or 0) or 0)
    job.base_discount_percent = float(request.form.get('base_discount_percent', job.base_discount_percent or 0) or 0)
    job.agent_discount_percent = float(request.form.get('agent_discount_percent', job.agent_discount_percent or 0) or 0)
    job.additional_discount_percent = float(request.form.get('additional_discount_percent', job.additional_discount_percent or 0) or 0)
    job.additional_charges = float(request.form.get('additional_charges', job.additional_charges or 0) or 0)
    job.final_price = float(request.form.get('final_price', job.final_price or 0) or 0)
    job.invoice_number = request.form.get('invoice_number', job.invoice_number or '').strip()
    
    db.session.commit()
    flash('Job updated successfully!', 'success')
    return redirect(url_for('view_job', job_id=job.id))


@app.route('/jobs/delete/<int:job_id>', methods=['POST'])
@login_required
def delete_job(job_id):
    try:
        job = Job.query.get_or_404(job_id)
        db.session.delete(job)
        db.session.commit()
        flash('Job deleted successfully!', 'success')
        return redirect(url_for('jobs'))
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error deleting job {job_id}: {str(e)}')
        flash('Error deleting job. Please try again.', 'error')
        return redirect(url_for('jobs'))


@app.route('/jobs/update_status/<int:job_id>', methods=['POST'])
@login_required
@validate_json_input
def update_job_status(job_id):
    try:
        job = Job.query.get_or_404(job_id)
        data = request.get_json()
        
        if not data or 'status' not in data:
            return jsonify({'success': False, 'message': 'Status is required'}), 400
        
        new_status = data['status']
        valid_statuses = ['Scheduled', 'In Progress', 'Completed', 'Cancelled', 'Failed', 'No Show']
        
        if new_status not in valid_statuses:
            return jsonify({'success': False, 'message': 'Invalid status'}), 400
        
        job.status = new_status
        db.session.commit()
        
        app.logger.info(f'Job {job_id} status updated to {new_status} by user {current_user.username}')
        
        return jsonify({
            'success': True, 
            'message': f'Job status updated to {new_status}',
            'new_status': new_status
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'Error updating job status {job_id}: {str(e)}')
        return jsonify({'success': False, 'message': 'Error updating job status'}), 500


@app.route('/download-report', methods=['GET'])
@login_required
def download_report():
    # Query all data
    drivers = Driver.query.all()
    agents = Agent.query.all()
    vehicles = Vehicle.query.all()
    services = Service.query.all()
    billings = Billing.query.all()
    discounts = Discount.query.all()
    jobs = Job.query.all()
    print(discounts)
    # Prepare DataFrames for each sheet
    drivers_df = pd.DataFrame([
        {
            'ID': getattr(d, 'id', ''),
            'Name': getattr(d, 'name', ''),
            'Phone': getattr(d, 'phone', ''),
        } for d in drivers
    ])

    # AGENTS SHEET
    agents_df = pd.DataFrame([
        {
            'ID': getattr(a, 'id', ''),
            'Name': getattr(a, 'name', ''),
            'Email': getattr(a, 'email', ''),
            'Mobile': getattr(a, 'mobile', ''),
            'Type': getattr(a, 'type', ''),
            'Status': getattr(a, 'status', ''),
            'Agent Discount %': getattr(a, 'agent_discount_percent', 0.0),
        } for a in agents
    ])

    # VEHICLES SHEET
    vehicles_df = pd.DataFrame([
        {
            'ID': getattr(v, 'id', ''),
            'Name': getattr(v, 'name', ''),
            'Number': getattr(v, 'number', ''),
            'Type': getattr(v, 'type', ''),
            'Status': getattr(v, 'status', ''),
        } for v in vehicles
    ])

    # SERVICES SHEET
    services_df = pd.DataFrame([
        {
            'ID': getattr(s, 'id', ''),
            'Name': getattr(s, 'name', ''),
            'Description': getattr(s, 'description', ''),
            'Status': getattr(s, 'status', ''),
            'Base Price': getattr(s, 'base_price', 0.0),
        } for s in services
    ])

    # BILLING SHEET
    billings_df = pd.DataFrame([
        {
            'ID': getattr(b, 'id', ''),
            'Job ID': getattr(b, 'job_id', ''),
            'Invoice Number': getattr(b, 'invoice_number', ''),
            'Invoice Date': getattr(b, 'invoice_date', '').strftime('%Y-%m-%d') if getattr(b, 'invoice_date', None) else '',
            'Due Date': getattr(b, 'due_date', '').strftime('%Y-%m-%d') if getattr(b, 'due_date', None) else '',
            'Base Price': getattr(b, 'base_price', 0.0),
            'Base Discount Amount': getattr(b, 'base_discount_amount', 0.0),
            'Agent Discount Amount': getattr(b, 'agent_discount_amount', 0.0),
            'Additional Discount Amount': getattr(b, 'additional_discount_amount', 0.0),
            'Additional Charges': getattr(b, 'additional_charges', 0.0),
            'Subtotal': getattr(b, 'subtotal', 0.0),
            'Tax Amount': getattr(b, 'tax_amount', 0.0),
            'Total Amount': getattr(b, 'total_amount', 0.0),
            'Payment Status': getattr(b, 'payment_status', ''),
            'Payment Date': getattr(b, 'payment_date', '').strftime('%Y-%m-%d') if getattr(b, 'payment_date', None) else '',
            'Payment Method': getattr(b, 'payment_method', ''),
            'Discount ID': getattr(b, 'discount_id', ''),
            'Notes': getattr(b, 'notes', ''),
            'Terms & Conditions': getattr(b, 'terms_conditions', ''),
        } for b in billings
    ])

    # JOBS SHEET (for monthly sheets)
    jobs_df = pd.DataFrame([
        {
            'ID': getattr(j, 'id', ''),
            'Customer Name': getattr(j, 'customer_name', ''),
            'Customer Email': getattr(j, 'customer_email', ''),
            'Customer Mobile': getattr(j, 'customer_mobile', ''),
            'Customer Reference': getattr(j, 'customer_reference', ''),
            'Passenger Name': getattr(j, 'passenger_name', ''),
            'Passenger Email': getattr(j, 'passenger_email', ''),
            'Passenger Mobile': getattr(j, 'passenger_mobile', ''),
            'Type of Service': getattr(j, 'type_of_service', ''),
            'Service ID': getattr(j, 'service_id', ''),
            'Pickup Date': getattr(j, 'pickup_date', ''),
            'Pickup Time': getattr(j, 'pickup_time', ''),
            'Pickup Location': getattr(j, 'pickup_location', ''),
            'Dropoff Location': getattr(j, 'dropoff_location', ''),
            'Vehicle Type': getattr(j, 'vehicle_type', ''),
            'Vehicle Number': getattr(j, 'vehicle_number', ''),
            'Driver Contact': getattr(j, 'driver_contact', ''),
            'Payment Mode': getattr(j, 'payment_mode', ''),
            'Payment Status': getattr(j, 'payment_status', ''),
            'Order Status': getattr(j, 'order_status', ''),
            'Message': getattr(j, 'message', ''),
            'Remarks': getattr(j, 'remarks', ''),
            'Has Additional Stop': getattr(j, 'has_additional_stop', False),
            'Additional Stops': getattr(j, 'additional_stops', ''),
            'Has Request': getattr(j, 'has_request', False),
            'Reference': getattr(j, 'reference', ''),
            'Status': getattr(j, 'status', ''),
            'Date': getattr(j, 'date', ''),
            'Driver ID': getattr(j, 'driver_id', ''),
            'Agent ID': getattr(j, 'agent_id', ''),
            'Base Price': getattr(j, 'base_price', 0.0),
            'Base Discount %': getattr(j, 'base_discount_percent', 0.0),
            'Agent Discount %': getattr(j, 'agent_discount_percent', 0.0),
            'Additional Discount %': getattr(j, 'additional_discount_percent', 0.0),
            'Additional Charges': getattr(j, 'additional_charges', 0.0),
            'Final Price': getattr(j, 'final_price', 0.0),
            'Invoice Number': getattr(j, 'invoice_number', ''),
        } for j in jobs
    ])
    # Parse pickup_date to datetime for grouping
    if not jobs_df.empty:
        jobs_df['Pickup Date'] = pd.to_datetime(jobs_df['Pickup Date'], errors='coerce')
        jobs_df['Month'] = jobs_df['Pickup Date'].dt.month
        jobs_df['Year'] = jobs_df['Pickup Date'].dt.year

    discounts_df = pd.DataFrame([
        {
            'ID': getattr(d, 'id', ''),
            'Name': getattr(d, 'name', ''),
            'Code': getattr(d, 'code', ''),
            'Percent': getattr(d, 'percent', 0.0),
            'Amount': getattr(d, 'amount', 0.0),
            'Discount Type': getattr(d, 'discount_type', ''),
            'Is Base Discount': getattr(d, 'is_base_discount', False),
            'Is Active': getattr(d, 'is_active', False),
            'Valid From': getattr(d, 'valid_from', '').strftime('%Y-%m-%d') if getattr(d, 'valid_from', None) else '',
            'Valid To': getattr(d, 'valid_to', '').strftime('%Y-%m-%d') if getattr(d, 'valid_to', None) else ''
        }
        for d in discounts
    ])

    # Write to Excel in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        drivers_df.to_excel(writer, sheet_name='Drivers', index=False)
        agents_df.to_excel(writer, sheet_name='Agents', index=False)
        vehicles_df.to_excel(writer, sheet_name='Vehicles', index=False)
        services_df.to_excel(writer, sheet_name='Services', index=False)
        billings_df.to_excel(writer, sheet_name='Billing', index=False)
        discounts_df.to_excel(writer, sheet_name='Discounts', index=False)
        # Monthly job sheets with unique names
        if not jobs_df.empty:
            for (year, month), group in jobs_df.groupby(['Year', 'Month']):
                if pd.isna(year) or pd.isna(month):
                    continue
                month_name = pd.Timestamp(year=int(year), month=int(month), day=1).strftime('%B')
                sheet_name = f'Jobs - {month_name} {int(year)}'
                group = group.drop(['Month', 'Year'], axis=1)
                group.to_excel(writer, sheet_name=sheet_name, index=False)
        # Formatting: bold headers, date formatting, autofit, borders, header fill
        for sheet_name, ws in writer.sheets.items():
            # Bold headers and header fill
            header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # Date formatting for Pickup Date
            headers = [cell.value for cell in ws[1]]
            if 'Pickup Date' in headers:
                col_index = headers.index('Pickup Date') + 1
                for row in ws.iter_rows(min_row=2, min_col=col_index, max_col=col_index):
                    for cell in row:
                        cell.number_format = 'YYYY-MM-DD'
            # Autofit column widths
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ''
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2
            # Add borders to all cells
            thin = Side(border_style="thin", color="B7B7B7")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='fleet_report.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/jobs/download', methods=['POST'])
@login_required
def download_jobs():
    import csv
    from io import StringIO
    from datetime import datetime
    
    selected_jobs = request.form.getlist('selected_jobs')
    
    if not selected_jobs:
        flash('No jobs selected for download', 'error')
        return redirect(url_for('jobs'))
    
    # Get the selected jobs
    jobs = Job.query.filter(Job.id.in_(selected_jobs)).all()
    
    # Create CSV data
    output = StringIO()
    writer = csv.writer(output)
    
    # Write header
    writer.writerow([
        'Job ID', 'Customer Name', 'Customer Email', 'Customer Mobile', 'Customer Reference',
        'Passenger Name', 'Passenger Email', 'Passenger Mobile', 'Type of Service',
        'Pickup Date', 'Pickup Time', 'Pickup Location', 'Drop-off Location',
        'Vehicle Type', 'Vehicle Number', 'Driver Contact', 'Driver ID',
        'Payment Mode', 'Payment Status', 'Order Status', 'Message', 'Remarks',
        'Reference', 'Status', 'Date'
    ])
    
    # Write data rows
    for job in jobs:
        writer.writerow([
            job.id,
            job.customer_name or '',
            job.customer_email or '',
            job.customer_mobile or '',
            job.customer_reference or '',
            job.passenger_name or '',
            job.passenger_email or '',
            job.passenger_mobile or '',
            job.type_of_service or '',
            job.pickup_date or '',
            job.pickup_time or '',
            job.pickup_location or '',
            job.dropoff_location or '',
            job.vehicle_type or '',
            job.vehicle_number or '',
            job.driver_contact or '',
            job.driver_id or '',
            job.payment_mode or '',
            job.payment_status or '',
            job.order_status or '',
            job.message or '',
            job.remarks or '',
            job.reference or '',
            job.status or '',
            job.date or ''
        ])
    
    # Create response
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=jobs_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    
    return response


@app.route('/jobs/smart_add', methods=['GET', 'POST'])
@login_required
def smart_add_job():
    parsed_data = None
    if request.method == 'POST':
        message = request.form.get('message')
        parsed_data = parse_job_message(message)
        return render_template('view_job.html', job=parsed_data, smart_add=True, pasted_message=message)
    return render_template('smart_add.html')


def parse_job_message(message):
    # Try to parse as field: value pairs
    data = {}
    lines = message.splitlines()
    for line in lines:
        if ':' in line:
            field, value = line.split(':', 1)
            field = field.strip().lower().replace(' ', '_')
            value = value.strip()
            # Map common aliases to job fields
            field_map = {
                'agent': 'customer_name',
                'agent_email': 'customer_email',
                'agent_mobile': 'customer_mobile',
                'service': 'type_of_service',
                'vehicle': 'vehicle_type',
                'vehicle_number': 'vehicle_number',
                'pickup': 'pickup_location',
                'drop': 'dropoff_location',
                'date': 'pickup_date',
                'time': 'pickup_time',
                'status': 'status',
                'passenger': 'passenger_name',
                'passenger_email': 'passenger_email',
                'passenger_mobile': 'passenger_mobile',
                'reference': 'reference',
                'remarks': 'remarks',
                'message': 'message',
            }
            mapped_field = field_map.get(field, field)
            data[mapped_field] = value
    # Fallback to regex for common fields if not found
    if not data:
        patterns = {
            'customer_name': r'Customer[:\-]?\s*([\w\s]+)',
            'customer_email': r'Email[:\-]?\s*([\w\.-]+@[\w\.-]+)',
            'customer_mobile': r'Mobile[:\-]?\s*(\d+)',
            'type_of_service': r'Service[:\-]?\s*([\w\s]+)',
            'pickup_date': r'Date[:\-]?\s*([\d\-/]+)',
            'pickup_time': r'Time[:\-]?\s*([\d:apmAPM\s]+)',
            'pickup_location': r'Pickup[:\-]?\s*([\w\s]+)',
            'dropoff_location': r'Drop[:\-]?\s*([\w\s]+)',
            'vehicle_type': r'Vehicle[:\-]?\s*([\w\s]+)',
            'driver_contact': r'Driver[:\-]?\s*([\w\s]+)',
            'payment_status': r'Payment[:\-]?\s*([\w\s]+)',
            'order_status': r'Status[:\-]?\s*([\w\s]+)',
        }
        for field, pattern in patterns.items():
            match = re.search(pattern, message, re.IGNORECASE)
            if match:
                data[field] = match.group(1).strip()
    return data


# DRIVERS CRUD
@app.route('/drivers')
@login_required
def drivers():
    name = request.args.get('name', '')
    phone = request.args.get('phone', '')
    query = Driver.query
    if name:
        query = query.filter(Driver.name.ilike(f'%{name}%'))
    if phone:
        query = query.filter(Driver.phone.ilike(f'%{phone}%'))
    drivers = query.all()
    if request.headers.get('HX-Request') == 'true':
        return render_template('drivers_table.html', drivers=drivers)
    return render_template('drivers.html', drivers=drivers)


@app.route('/drivers/add', methods=['GET', 'POST'])
@login_required
def add_driver():
    errors = {}
    if request.method == 'POST':
        name = request.form['name']
        phone = request.form['phone']
        driver = Driver(name=name, phone=phone)
        db.session.add(driver)
        db.session.commit()
        if request.headers.get('HX-Request') == 'true':
            drivers = Driver.query.all()
            response = make_response(render_template('drivers_table.html', drivers=drivers))
            response.headers['HX-Trigger'] = 'closeModal'
            return response
        return redirect(url_for('drivers'))
    return render_template('driver_form.html', action='Add', driver=None, errors=errors,
                           action_url=url_for('add_driver'), hx_post_url=url_for('add_driver'),
                           hx_target='#drivers-table', hx_swap='outerHTML')


@app.route('/drivers/edit/<int:driver_id>', methods=['GET', 'POST'])
@login_required
def edit_driver(driver_id):
    driver = Driver.query.get_or_404(driver_id)
    errors = {}
    if request.method == 'POST':
        driver.name = request.form['name']
        driver.phone = request.form['phone']
        db.session.commit()
        return redirect(url_for('drivers'))
    if request.headers.get('HX-Request') == 'true':
        return render_template('driver_form.html', action='Edit', driver=driver, errors=errors,
                               action_url=url_for('edit_driver', driver_id=driver.id),
                               hx_post_url=url_for('edit_driver', driver_id=driver.id), hx_target='#drivers-table',
                               hx_swap='outerHTML')
    return render_template('edit_driver_page.html', action='Edit', driver=driver, errors=errors,
                           action_url=url_for('edit_driver', driver_id=driver.id), hx_post_url=None, hx_target=None,
                           hx_swap=None)


@app.route('/drivers/delete/<int:driver_id>', methods=['POST'])
@login_required
def delete_driver(driver_id):
    driver = Driver.query.get_or_404(driver_id)
    db.session.delete(driver)
    db.session.commit()
    return redirect(url_for('drivers'))


# AGENTS CRUD
@app.route('/agents')
@login_required
def agents():
    name = request.args.get('name', '')
    email = request.args.get('email', '')
    mobile = request.args.get('mobile', '')
    type_ = request.args.get('type', '')
    status = request.args.get('status', '')
    query = Agent.query
    if name:
        query = query.filter(Agent.name.ilike(f'%{name}%'))
    if email:
        query = query.filter(Agent.email.ilike(f'%{email}%'))
    if mobile:
        query = query.filter(Agent.mobile.ilike(f'%{mobile}%'))
    if type_:
        query = query.filter(Agent.type.ilike(f'%{type_}%'))
    if status:
        query = query.filter(Agent.status == status)
    agents = query.all()
    if request.headers.get('HX-Request') == 'true':
        return render_template('agents_table.html', agents=agents)
    return render_template('agents.html', agents=agents)


@app.route('/agents/add', methods=['GET', 'POST'])
@login_required
def add_agent():
    errors = {}
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        mobile = request.form['mobile']
        type_ = request.form['type']
        status = request.form['status']
        agent_discount_percent = float(request.form.get('agent_discount_percent', 0))
        agent = Agent(name=name, email=email, mobile=mobile, type=type_, status=status, agent_discount_percent=agent_discount_percent)
        db.session.add(agent)
        db.session.commit()
        if request.headers.get('HX-Request') == 'true':
            agents = Agent.query.all()
            response = make_response(render_template('agents_table.html', agents=agents))
            response.headers['HX-Trigger'] = 'closeModal'
            return response
        return redirect(url_for('agents'))
    # Serve only the form partial for HTMX/modal requests
    if request.headers.get('HX-Request') == 'true':
        return render_template('agent_form.html', action='Add', agent=None, errors=errors,
                               action_url=url_for('add_agent'), hx_post_url=url_for('add_agent'),
                               hx_target='#agents-table', hx_swap='outerHTML')
    return render_template('add_agent.html', action='Add', agent=None, errors=errors)


@app.route('/agents/add_ajax', methods=['POST'])
@login_required
def add_agent_ajax():
    errors = {}
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip()
    mobile = request.form.get('mobile', '').strip()
    type_ = request.form.get('type', '').strip()
    status = request.form.get('status', 'Active').strip()
    if not name:
        errors['name'] = ['Name is required.']
    # Optionally add more validation here
    if errors:
        # Render the form with errors for HTMX swap
        return render_template('agent_form.html', action='Add Agent', agent=None, errors=errors,
                               action_url=url_for('add_agent_ajax'), hx_post_url=url_for('add_agent_ajax'), hx_target='#agent-modal-body', hx_swap='outerHTML')
    agent_discount_percent = float(request.form.get('agent_discount_percent', 0))
    agent = Agent(name=name, email=email, mobile=mobile, type=type_, status=status, agent_discount_percent=agent_discount_percent)
    db.session.add(agent)
    db.session.commit()
    # Return JSON for JS to update dropdown and close modal
    return jsonify({'success': True, 'id': agent.id, 'name': agent.name})


@app.route('/agents/edit/<int:agent_id>', methods=['GET', 'POST'])
@login_required
def edit_agent(agent_id):
    agent = Agent.query.get_or_404(agent_id)
    errors = {}
    if request.method == 'POST':
        agent.name = request.form['name']
        agent.email = request.form['email']
        agent.mobile = request.form['mobile']
        agent.type = request.form['type']
        agent.status = request.form['status']
        agent.agent_discount_percent = float(request.form.get('agent_discount_percent', 0))
        db.session.commit()
        return redirect(url_for('agents'))
    if request.headers.get('HX-Request') == 'true':
        return render_template('agent_form.html', action='Edit', agent=agent, errors=errors,
                               action_url=url_for('edit_agent', agent_id=agent.id),
                               hx_post_url=url_for('edit_agent', agent_id=agent.id), hx_target='#agents-table',
                               hx_swap='outerHTML')
    return render_template('edit_agent_page.html', action='Edit', agent=agent, errors=errors,
                           action_url=url_for('edit_agent', agent_id=agent.id), hx_post_url=None, hx_target=None,
                           hx_swap=None)


@app.route('/agents/delete/<int:agent_id>', methods=['POST'])
@login_required
def delete_agent(agent_id):
    agent = Agent.query.get_or_404(agent_id)
    db.session.delete(agent)
    db.session.commit()
    return redirect(url_for('agents'))


# BILLING CRUD
@app.route('/billing')
@login_required
def billing():
    billings = Billing.query.all()
    return render_template('billing.html', billings=billings)


@app.route('/billing/add', methods=['GET', 'POST'])
@login_required
def add_billing():
    from models import Job
    jobs = Job.query.all()
    
    if request.method == 'POST':
        try:
            # Get the selected job
            job_id = request.form['job_id']
            job = Job.query.get(job_id)
            
            if not job:
                flash('Selected job not found', 'error')
                return render_template('billing_form.html', action='Add', jobs=jobs)
            
            # Create billing record with all the new fields
            base_price = job.base_price or 0
            base_discount_amount = job.base_discount_percent * base_price / 100 if job.base_discount_percent else 0
            agent_discount_amount = job.agent_discount_percent * base_price / 100 if job.agent_discount_percent else 0
            additional_discount_amount = job.additional_discount_percent * base_price / 100 if job.additional_discount_percent else 0
            additional_charges = float(request.form.get('additional_charges', 0))
            subtotal = base_price - base_discount_amount - agent_discount_amount - additional_discount_amount
            tax_amount = float(request.form.get('tax_amount', 0))
            total_amount = subtotal + additional_charges + tax_amount
            
            billing = Billing(
                job_id=job_id,
                invoice_number=request.form.get('invoice_number') or f'INV-{job_id}-{datetime.now().strftime("%Y%m%d%H%M%S")}',
                invoice_date=request.form.get('invoice_date'),
                due_date=request.form.get('due_date'),
                base_price=base_price,
                base_discount_amount=base_discount_amount,
                agent_discount_amount=agent_discount_amount,
                additional_discount_amount=additional_discount_amount,
                additional_charges=additional_charges,
                subtotal=subtotal,
                tax_amount=tax_amount,
                total_amount=total_amount,
                payment_status=request.form.get('payment_status', 'Pending'),
                payment_date=request.form.get('payment_date'),
                payment_method=request.form.get('payment_method'),
                notes=request.form.get('notes'),
                terms_conditions=request.form.get('terms_conditions')
            )
            
            db.session.add(billing)
            db.session.commit()
            flash('Billing record created successfully', 'success')
            return redirect(url_for('billing'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating billing record: {str(e)}', 'error')
            return render_template('billing_form.html', action='Add', jobs=jobs)
    
    return render_template('billing_form.html', action='Add', jobs=jobs)


@app.route('/billing/edit/<int:billing_id>', methods=['GET', 'POST'])
@login_required
def edit_billing(billing_id):
    from models import Job
    billing = Billing.query.get_or_404(billing_id)
    jobs = Job.query.all()
    
    if request.method == 'POST':
        try:
            # Get the selected job
            job_id = request.form['job_id']
            job = Job.query.get(job_id)
            
            if not job:
                flash('Selected job not found', 'error')
                return render_template('billing_form.html', action='Edit', billing=billing, jobs=jobs)
            
            # Update billing record with all the new fields
            base_price = job.base_price or 0
            base_discount_amount = job.base_discount_percent * base_price / 100 if job.base_discount_percent else 0
            agent_discount_amount = job.agent_discount_percent * base_price / 100 if job.agent_discount_percent else 0
            additional_discount_amount = job.additional_discount_percent * base_price / 100 if job.additional_discount_percent else 0
            additional_charges = float(request.form.get('additional_charges', 0))
            subtotal = base_price - base_discount_amount - agent_discount_amount - additional_discount_amount
            tax_amount = float(request.form.get('tax_amount', 0))
            total_amount = subtotal + additional_charges + tax_amount
            
            billing.job_id = job_id
            billing.invoice_number = request.form.get('invoice_number') or billing.invoice_number or f'INV-{job_id}-{datetime.now().strftime("%Y%m%d%H%M%S")}'
            billing.invoice_date = request.form.get('invoice_date')
            billing.due_date = request.form.get('due_date')
            billing.base_price = base_price
            billing.base_discount_amount = base_discount_amount
            billing.agent_discount_amount = agent_discount_amount
            billing.additional_discount_amount = additional_discount_amount
            billing.additional_charges = additional_charges
            billing.subtotal = subtotal
            billing.tax_amount = tax_amount
            billing.total_amount = total_amount
            billing.payment_status = request.form.get('payment_status', 'Pending')
            billing.payment_date = request.form.get('payment_date')
            billing.payment_method = request.form.get('payment_method')
            billing.notes = request.form.get('notes')
            billing.terms_conditions = request.form.get('terms_conditions')
            
            db.session.commit()
            flash('Billing record updated successfully', 'success')
            return redirect(url_for('billing'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating billing record: {str(e)}', 'error')
            return render_template('billing_form.html', action='Edit', billing=billing, jobs=jobs)
    
    return render_template('billing_form.html', action='Edit', billing=billing, jobs=jobs)


@app.route('/billing/delete/<int:billing_id>', methods=['POST'])
@login_required
def delete_billing(billing_id):
    billing = Billing.query.get_or_404(billing_id)
    db.session.delete(billing)
    db.session.commit()
    return redirect(url_for('billing'))


# DISCOUNTS CRUD
@app.route('/discounts')
@login_required
def discounts():
    discounts = Discount.query.all()
    return render_template('discounts.html', discounts=discounts)


@app.route('/discounts/add', methods=['GET', 'POST'])
@login_required
def add_discount():
    if request.method == 'POST':
        code = request.form['code']
        percent = request.form['percent']
        discount = Discount(code=code, percent=percent)
        db.session.add(discount)
        db.session.commit()
        return redirect(url_for('discounts'))
    return render_template('discount_form.html', action='Add')


@app.route('/discounts/edit/<int:discount_id>', methods=['GET', 'POST'])
@login_required
def edit_discount(discount_id):
    discount = Discount.query.get_or_404(discount_id)
    if request.method == 'POST':
        discount.code = request.form['code']
        discount.percent = request.form['percent']
        db.session.commit()
        return redirect(url_for('discounts'))
    return render_template('discount_form.html', action='Edit', discount=discount)


@app.route('/discounts/delete/<int:discount_id>', methods=['POST'])
@login_required
def delete_discount(discount_id):
    discount = Discount.query.get_or_404(discount_id)
    db.session.delete(discount)
    db.session.commit()
    return redirect(url_for('discounts'))


# SERVICES CRUD
@app.route('/services')
@login_required
def services():
    name = request.args.get('name', '')
    status = request.args.get('status', '')
    query = Service.query
    if name:
        query = query.filter(Service.name.ilike(f'%{name}%'))
    if status:
        query = query.filter(Service.status == status)
    services = query.all()
    if request.headers.get('HX-Request') == 'true':
        return render_template('services.html', services=services)
    return render_template('services.html', services=services)


@app.route('/services/add', methods=['GET', 'POST'])
@login_required
def add_service():
    if request.method == 'POST':
        name = request.form['name']
        description = request.form['description']
        status = request.form['status']
        base_price = float(request.form.get('base_price', 0))
        service = Service(name=name, description=description, status=status, base_price=base_price)
        db.session.add(service)
        db.session.commit()
        return redirect(url_for('services'))
    return render_template('add_service.html', action='Add', service=None)


@app.route('/services/edit/<int:service_id>', methods=['GET', 'POST'])
@login_required
def edit_service(service_id):
    service = Service.query.get_or_404(service_id)
    if request.method == 'POST':
        service.name = request.form['name']
        service.description = request.form['description']
        service.status = request.form['status']
        service.base_price = float(request.form.get('base_price', 0))
        db.session.commit()
        return redirect(url_for('services'))
    return render_template('edit_service.html', action='Edit', service=service)


@app.route('/services/delete/<int:service_id>', methods=['POST'])
@login_required
def delete_service(service_id):
    service = Service.query.get_or_404(service_id)
    db.session.delete(service)
    db.session.commit()
    return redirect(url_for('services'))


@app.route('/services/add_ajax', methods=['POST'])
@login_required
def add_service_ajax():
    errors = {}
    name = request.form.get('name', '').strip()
    description = request.form.get('description', '').strip()
    status = request.form.get('status', 'Active').strip()
    base_price = float(request.form.get('base_price', 0))
    if not name:
        errors['name'] = ['Name is required.']
    if errors:
        return render_template('service_form.html', action='Add Service', service=None, errors=errors,
                               action_url=url_for('add_service_ajax'), hx_post_url=url_for('add_service_ajax'), hx_target='#service-modal-body', hx_swap='outerHTML')
    service = Service(name=name, description=description, status=status, base_price=base_price)
    db.session.add(service)
    db.session.commit()
    return jsonify({'success': True, 'id': service.id, 'name': service.name})


# VEHICLES CRUD
@app.route('/vehicles')
@login_required
def vehicles():
    name = request.args.get('name', '')
    number = request.args.get('number', '')
    type_ = request.args.get('type', '')
    status = request.args.get('status', '')
    query = Vehicle.query
    if name:
        query = query.filter(Vehicle.name.ilike(f'%{name}%'))
    if number:
        query = query.filter(Vehicle.number.ilike(f'%{number}%'))
    if type_:
        query = query.filter(Vehicle.type.ilike(f'%{type_}%'))
    if status:
        query = query.filter(Vehicle.status == status)
    vehicles = query.all()
    if request.headers.get('HX-Request') == 'true':
        return render_template('vehicles.html', vehicles=vehicles)
    return render_template('vehicles.html', vehicles=vehicles)


@app.route('/vehicles/add', methods=['GET', 'POST'])
@login_required
def add_vehicle():
    errors = {}
    if request.method == 'POST':
        name = request.form['name']
        number = request.form['number']
        type_ = request.form['type']
        status = request.form['status']
        vehicle = Vehicle(name=name, number=number, type=type_, status=status)
        db.session.add(vehicle)
        db.session.commit()
        return redirect(url_for('vehicles'))
    return render_template('add_vehicle.html', action='Add', vehicle=None, errors=errors)


@app.route('/vehicles/edit/<int:vehicle_id>', methods=['GET', 'POST'])
@login_required
def edit_vehicle(vehicle_id):
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    errors = {}
    if request.method == 'POST':
        vehicle.name = request.form['name']
        vehicle.number = request.form['number']
        vehicle.type = request.form['type']
        vehicle.status = request.form['status']
        db.session.commit()
        return redirect(url_for('vehicles'))
    return render_template('edit_vehicle.html', action='Edit', vehicle=vehicle, errors=errors)


@app.route('/vehicles/delete/<int:vehicle_id>', methods=['POST'])
@login_required
def delete_vehicle(vehicle_id):
    vehicle = Vehicle.query.get_or_404(vehicle_id)
    db.session.delete(vehicle)
    db.session.commit()
    return redirect(url_for('vehicles'))


@app.route('/vehicles/add_ajax', methods=['POST'])
@login_required
def add_vehicle_ajax():
    errors = {}
    name = request.form.get('name', '').strip()
    number = request.form.get('number', '').strip()
    type_ = request.form.get('type', '').strip()
    status = request.form.get('status', 'Active').strip()
    if not name:
        errors['name'] = ['Name is required.']
    if not number:
        errors['number'] = ['Number is required.']
    if errors:
        return render_template('vehicle_form.html', action='Add Vehicle', vehicle=None, errors=errors,
                               action_url=url_for('add_vehicle_ajax'), hx_post_url=url_for('add_vehicle_ajax'), hx_target='#vehicle-modal-body', hx_swap='outerHTML')
    vehicle = Vehicle(name=name, number=number, type=type_, status=status)
    db.session.add(vehicle)
    db.session.commit()
    return jsonify({'success': True, 'id': vehicle.id, 'name': f'{vehicle.number} ({vehicle.name})'})

@app.route('/drivers/add_ajax', methods=['POST'])
@login_required
def add_driver_ajax():
    errors = {}
    name = request.form.get('name', '').strip()
    phone = request.form.get('phone', '').strip()
    if not name:
        errors['name'] = ['Name is required.']
    if not phone:
        errors['phone'] = ['Phone is required.']
    if errors:
        return render_template('driver_form.html', action='Add Driver', driver=None, errors=errors,
                               action_url=url_for('add_driver_ajax'), hx_post_url=url_for('add_driver_ajax'), hx_target='#driver-modal-body', hx_swap='outerHTML')
    driver = Driver(name=name, phone=phone)
    db.session.add(driver)
    db.session.commit()
    return jsonify({'success': True, 'id': driver.id, 'name': f'{driver.name} ({driver.phone})'})


@app.route('/api/calculate_pricing', methods=['POST'])
@login_required
def calculate_pricing():
    """Calculate pricing for a service and agent combination"""
    try:
        data = request.get_json()
        service_id = data.get('service_id')
        agent_id = data.get('agent_id')
        additional_discount = float(data.get('additional_discount_percent', 0))
        additional_charges = float(data.get('additional_charges', 0))
        
        if not service_id or not agent_id:
            return jsonify({'success': False, 'error': 'Service and agent are required'})
        
        # Get service base price
        service = Service.query.get(service_id)
        if not service:
            return jsonify({'success': False, 'error': 'Service not found'})
        
        base_price = service.base_price
        
        # Get base discount (system-wide discount)
        base_discount = Discount.query.filter_by(is_base_discount=True, is_active=True).first()
        base_discount_percent = (base_discount.percent if base_discount else 0.0) or 0.0
        
        # Get agent discount
        agent = Agent.query.get(agent_id)
        agent_discount_percent = (agent.agent_discount_percent if agent else 0.0) or 0.0
        
        # Calculate discount amounts
        base_discount_amount = (base_price * base_discount_percent) / 100
        agent_discount_amount = (base_price * agent_discount_percent) / 100
        additional_discount_amount = (base_price * additional_discount) / 100
        
        # Calculate final price
        subtotal = base_price - base_discount_amount - agent_discount_amount - additional_discount_amount
        final_price = subtotal + additional_charges
        
        return jsonify({
            'success': True,
            'pricing': {
                'base_price': base_price,
                'base_discount_percent': base_discount_percent,
                'base_discount_amount': base_discount_amount,
                'agent_discount_percent': agent_discount_percent,
                'agent_discount_amount': agent_discount_amount,
                'additional_discount_percent': additional_discount,
                'additional_discount_amount': additional_discount_amount,
                'additional_charges': additional_charges,
                'subtotal': subtotal,
                'final_price': final_price
            }
        })
    except Exception as e:
        app.logger.error(f'Error calculating pricing: {str(e)}')
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/invoice/<int:billing_id>', methods=['GET'])
@login_required
def get_invoice(billing_id):
    """Get invoice details for modal display"""
    try:
        billing = Billing.query.get_or_404(billing_id)
        html = render_template('invoice_details.html', billing=billing)
        return jsonify({'success': True, 'html': html})
    except Exception as e:
        app.logger.error(f'Error getting invoice: {str(e)}')
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/invoice/<int:billing_id>/pdf', methods=['GET'])
@login_required
def download_invoice_pdf(billing_id):
    """Download invoice as PDF"""
    try:
        billing = Billing.query.get_or_404(billing_id)
        # This would generate actual PDF using a library like reportlab or weasyprint
        # For now, return a simple text response
        response = make_response(f"""
        INVOICE
        
        Invoice Number: {billing.invoice_number or 'N/A'}
        Date: {billing.invoice_date or 'N/A'}
        
        Job Details:
        - From: {billing.job.pickup_location or 'N/A'}
        - To: {billing.job.dropoff_location or 'N/A'}
        - Date: {billing.job.pickup_date or 'N/A'}
        - Time: {billing.job.pickup_time or 'N/A'}
        
        Pricing:
        - Base Price: SGD {(billing.base_price or 0):.2f}
        - Base Discount: SGD {(billing.base_discount_amount or 0):.2f}
        - Agent Discount: SGD {(billing.agent_discount_amount or 0):.2f}
        - Additional Discount: SGD {(billing.additional_discount_amount or 0):.2f}
        - Additional Charges: SGD {(billing.additional_charges or 0):.2f}
        - Total: SGD {(billing.total_amount or 0):.2f}
        """)
        response.headers['Content-Type'] = 'text/plain'
        response.headers['Content-Disposition'] = f'attachment; filename=invoice_{billing.invoice_number}.txt'
        return response
    except Exception as e:
        app.logger.error(f'Error generating PDF: {str(e)}')
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/billing/report/pdf', methods=['GET'])
@login_required
def generate_billing_report_pdf():
    """Generate PDF report of all invoices"""
    try:
        billings = Billing.query.all()
        # This would generate actual PDF using a library like reportlab or weasyprint
        # For now, return a simple text response
        report_content = "BILLING REPORT\n\n"
        for billing in billings:
            report_content += f"""
            Invoice: {billing.invoice_number or 'N/A'}
            Amount: SGD {(billing.total_amount or 0):.2f}
            Status: {billing.payment_status or 'N/A'}
            Date: {billing.invoice_date or 'N/A'}
            ---
            """
        
        response = make_response(report_content)
        response.headers['Content-Type'] = 'text/plain'
        response.headers['Content-Disposition'] = 'attachment; filename=billing_report.txt'
        return response
    except Exception as e:
        app.logger.error(f'Error generating billing report: {str(e)}')
        return jsonify({'success': False, 'error': str(e)})


# Import API routes
import api_routes

# Chat API Routes
@app.route('/api/chat', methods=['POST'])
@login_required
@csrf.exempt
def chat_api():
    try:
        data = request.get_json()
        message = data.get('message', '').lower().strip()
        
        # Parse the message and generate response
        response, data = parse_chat_message(message)
        
        return jsonify({
            'response': response,
            'data': data
        })
        
    except Exception as e:
        return jsonify({
            'response': f'Sorry, I encountered an error: {str(e)}',
            'data': None
        }), 500

@app.route('/api/chat/download', methods=['POST'])
@login_required
@csrf.exempt
def chat_download():
    try:
        data = request.get_json()
        query = data.get('query', '')
        table_data = data.get('data', [])
        
        if not table_data:
            return jsonify({'error': 'No data to download'}), 400
        
        # Create CSV content
        import csv
        import io
        
        output = io.StringIO()
        if table_data:
            writer = csv.DictWriter(output, fieldnames=table_data[0].keys())
            writer.writeheader()
            writer.writerows(table_data)
        
        csv_content = output.getvalue()
        output.close()
        
        # Create response
        response = make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename="{query.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        return response
        
    except Exception as e:
        return jsonify({'error': f'Download error: {str(e)}'}), 500

def parse_chat_message(message):
    """Parse chat message and return appropriate response and data"""
    
    # Jobs queries
    if re.search(r'\b(all\s+)?jobs?\b', message):
        return handle_jobs_query(message)
    
    # Driver queries
    elif re.search(r'\bdrivers?\b', message):
        return handle_drivers_query(message)
    
    # Vehicle queries
    elif re.search(r'\bvehicles?\b', message):
        return handle_vehicles_query(message)
    
    # Agent queries
    elif re.search(r'\bagents?\b', message):
        return handle_agents_query(message)
    
    # Service queries
    elif re.search(r'\bservices?\b', message):
        return handle_services_query(message)
    
    # Billing queries
    elif re.search(r'\bbilling?\b', message):
        return handle_billing_query(message)
    
    # Payment queries
    elif re.search(r'\bpayment\b', message):
        return handle_payment_query(message)
    
    # Status queries
    elif re.search(r'\bstatus\b', message):
        return handle_status_query(message)
    
    # Dashboard/Summary queries
    elif re.search(r'\b(dashboard|summary|overview)\b', message):
        return handle_dashboard_query(message)
    
    # Help
    elif re.search(r'\b(help|what can you do)\b', message):
        return handle_help_query(message)
    
    else:
        return "I'm not sure what you're asking for. Try asking about jobs, drivers, vehicles, agents, services, or payment status.", None

def handle_jobs_query(message):
    """Handle job-related queries"""
    
    if re.search(r'\bactive\b', message):
        jobs = Job.query.filter(Job.order_status.in_(['New', 'In Progress'])).limit(10).all()
        return f"I found {len(jobs)} active jobs:", [format_job(job) for job in jobs]
    
    elif re.search(r'\bpending\b', message):
        jobs = Job.query.filter(Job.order_status == 'Pending').limit(10).all()
        return f"I found {len(jobs)} pending jobs:", [format_job(job) for job in jobs]
    
    elif re.search(r'\bcompleted\b', message):
        jobs = Job.query.filter(Job.order_status == 'Completed').limit(10).all()
        return f"I found {len(jobs)} completed jobs:", [format_job(job) for job in jobs]
    
    elif re.search(r'\bcancelled\b', message):
        jobs = Job.query.filter(Job.order_status == 'Cancelled').limit(10).all()
        return f"I found {len(jobs)} cancelled jobs:", [format_job(job) for job in jobs]
    
    elif re.search(r'\bunpaid\b', message):
        jobs = Job.query.filter(Job.payment_status == 'Unpaid').limit(10).all()
        return f"I found {len(jobs)} unpaid jobs:", [format_job(job) for job in jobs]
    
    elif re.search(r'\bpaid\b', message):
        jobs = Job.query.filter(Job.payment_status == 'Paid').limit(10).all()
        return f"I found {len(jobs)} paid jobs:", [format_job(job) for job in jobs]
    
    else:
        # All jobs
        jobs = Job.query.order_by(Job.id.desc()).limit(10).all()
        return f"I found {len(jobs)} recent jobs:", [format_job(job) for job in jobs]

def handle_drivers_query(message):
    """Handle driver-related queries"""
    
    if re.search(r'\bavailable\b', message):
        # Drivers not assigned to active jobs
        active_driver_ids = db.session.query(Job.driver_id).filter(
            Job.order_status.in_(['New', 'In Progress'])
        ).distinct().all()
        active_ids = [id[0] for id in active_driver_ids if id[0]]
        
        drivers = Driver.query.filter(~Driver.id.in_(active_ids)).all()
        return f"I found {len(drivers)} available drivers:", [format_driver(driver) for driver in drivers]
    
    else:
        drivers = Driver.query.limit(10).all()
        return f"I found {len(drivers)} drivers:", [format_driver(driver) for driver in drivers]

def handle_vehicles_query(message):
    """Handle vehicle-related queries"""
    
    if re.search(r'\bavailable\b', message):
        # Get vehicles that are not currently assigned to active jobs
        # Since Job model doesn't have vehicle_id, we'll check by vehicle number
        active_jobs = Job.query.filter(Job.order_status.in_(['New', 'In Progress'])).all()
        active_vehicle_numbers = [job.vehicle_number for job in active_jobs if job.vehicle_number]
        
        # Get vehicles not in active jobs
        available_vehicles = Vehicle.query.filter(~Vehicle.number.in_(active_vehicle_numbers)).all()
        return f"I found {len(available_vehicles)} available vehicles:", [format_vehicle(vehicle) for vehicle in available_vehicles]
    
    else:
        vehicles = Vehicle.query.limit(10).all()
        return f"I found {len(vehicles)} vehicles:", [format_vehicle(vehicle) for vehicle in vehicles]

def handle_agents_query(message):
    """Handle agent-related queries"""
    agents = Agent.query.limit(10).all()
    return f"I found {len(agents)} agents:", [format_agent(agent) for agent in agents]

def handle_services_query(message):
    """Handle service-related queries"""
    services = Service.query.limit(10).all()
    return f"I found {len(services)} services:", [format_service(service) for service in services]

def handle_billing_query(message):
    """Handle billing-related queries"""
    billings = Billing.query.limit(10).all()
    return f"I found {len(billings)} billing records:", [format_billing(billing) for billing in billings]

def handle_payment_query(message):
    """Handle payment-related queries"""
    
    if re.search(r'\bunpaid\b', message):
        jobs = Job.query.filter(Job.payment_status == 'Unpaid').limit(10).all()
        return f"I found {len(jobs)} unpaid jobs:", [format_job(job) for job in jobs]
    
    elif re.search(r'\bpaid\b', message):
        jobs = Job.query.filter(Job.payment_status == 'Paid').limit(10).all()
        return f"I found {len(jobs)} paid jobs:", [format_job(job) for job in jobs]
    
    else:
        # Payment summary
        total_jobs = Job.query.count()
        paid_jobs = Job.query.filter(Job.payment_status == 'Paid').count()
        unpaid_jobs = Job.query.filter(Job.payment_status == 'Unpaid').count()
        
        return f"Payment Summary:\n- Total Jobs: {total_jobs}\n- Paid: {paid_jobs}\n- Unpaid: {unpaid_jobs}", None

def handle_status_query(message):
    """Handle status-related queries"""
    
    # Job status summary
    new_jobs = Job.query.filter(Job.order_status == 'New').count()
    in_progress_jobs = Job.query.filter(Job.order_status == 'In Progress').count()
    completed_jobs = Job.query.filter(Job.order_status == 'Completed').count()
    cancelled_jobs = Job.query.filter(Job.order_status == 'Cancelled').count()
    
    return f"Job Status Summary:\n- New: {new_jobs}\n- In Progress: {in_progress_jobs}\n- Completed: {completed_jobs}\n- Cancelled: {cancelled_jobs}", None

def handle_dashboard_query(message):
    """Handle dashboard/summary queries"""
    
    # Overall summary
    total_jobs = Job.query.count()
    total_drivers = Driver.query.count()
    total_vehicles = Vehicle.query.count()
    total_agents = Agent.query.count()
    
    active_jobs = Job.query.filter(Job.order_status.in_(['New', 'In Progress'])).count()
    completed_jobs = Job.query.filter(Job.order_status == 'Completed').count()
    unpaid_jobs = Job.query.filter(Job.payment_status == 'Unpaid').count()
    
    return f"Fleet Dashboard Summary:\n- Total Jobs: {total_jobs}\n- Active Jobs: {active_jobs}\n- Completed Jobs: {completed_jobs}\n- Unpaid Jobs: {unpaid_jobs}\n- Total Drivers: {total_drivers}\n- Total Vehicles: {total_vehicles}\n- Total Agents: {total_agents}", None

def handle_help_query(message):
    """Handle help queries"""
    return """I can help you with the following queries:

**Jobs:**
- "Show all jobs"
- "Active jobs"
- "Pending jobs"
- "Completed jobs"
- "Unpaid jobs"

**Drivers:**
- "All drivers"
- "Available drivers"

**Vehicles:**
- "All vehicles"
- "Available vehicles"

**Others:**
- "Payment status"
- "Job status"
- "Dashboard summary"

Try asking me about any of these topics!""", None

# Data formatting functions
def format_job(job):
    return {
        'id': job.id,
        'customer_name': job.customer_name,
        'pickup_location': job.pickup_location,
        'dropoff_location': job.dropoff_location,
        'order_status': job.order_status,
        'payment_status': job.payment_status,
        'pickup_date': job.pickup_date,
        'driver_contact': job.driver_contact,
        'vehicle_type': job.vehicle_type
    }

def format_driver(driver):
    return {
        'id': driver.id,
        'name': driver.name,
        'phone': driver.phone
    }

def format_vehicle(vehicle):
    return {
        'id': vehicle.id,
        'name': vehicle.name,
        'number': vehicle.number,
        'type': vehicle.type,
        'status': vehicle.status
    }

def format_agent(agent):
    return {
        'id': agent.id,
        'name': agent.name,
        'email': agent.email,
        'mobile': agent.mobile,
        'type': agent.type,
        'status': agent.status
    }

def format_service(service):
    return {
        'id': service.id,
        'name': service.name,
        'description': service.description,
        'status': service.status
    }

def format_billing(billing):
    return {
        'id': billing.id,
        'job_id': billing.job_id,
        'invoice_number': billing.invoice_number,
        'total_amount': billing.total_amount,
        'payment_status': billing.payment_status,
        'invoice_date': billing.invoice_date.strftime('%Y-%m-%d') if billing.invoice_date else 'N/A',
        'base_price': billing.base_price,
        'discount_id': billing.discount_id
    }


@app.cli.command('create-admin')
@click.argument('username')
@click.argument('email')
@click.argument('password')
@with_appcontext
def create_admin(username, email, password):
    from models import User, Role
    from app import db
    fleet_manager_role = Role.query.filter_by(name='fleet_manager').first()
    if not fleet_manager_role:
        fleet_manager_role = Role(name='fleet_manager', description='Fleet Manager')
        db.session.add(fleet_manager_role)
        db.session.commit()
    system_admin_role = Role.query.filter_by(name='system_admin').first()
    if not system_admin_role:
        system_admin_role = Role(name='system_admin', description='System Administrator')
        db.session.add(system_admin_role)
        db.session.commit()
    user = User.query.filter_by(username=username).first()
    if user:
        click.echo(f'User {username} already exists.')
        return
    user = User()
    user.username = username
    user.email = email
    user.active = True
    user.set_password(password)
    user.roles.append(fleet_manager_role)
    user.roles.append(system_admin_role)
    db.session.add(user)
    db.session.commit()
    click.echo(f'Admin user {username} created successfully.')


# CSRF token is automatically handled by Flask-WTF and Flask-Security

@app.context_processor
def inject_role_helpers():
    def has_role(role_name):
        return any(role.name == role_name for role in getattr(current_user, 'roles', []))

    def has_any_role(*role_names):
        return any(role.name in role_names for role in getattr(current_user, 'roles', []))

    return dict(has_role=has_role, has_any_role=has_any_role)


@app.context_processor
def inject_csrf_token():
    from flask_wtf.csrf import generate_csrf
    return dict(csrf_token=generate_csrf)




if __name__ == '__main__':
    import os

    port = int(os.environ.get('PORT', 5000))
    app.run(debug=True, host='0.0.0.0', port=port)
